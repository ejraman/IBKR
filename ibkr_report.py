#!/usr/bin/env python3
"""IBKR Wheel Strategy Report - Clean Version"""
import pandas as pd, io, re, sys, os, argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

parser = argparse.ArgumentParser()
parser.add_argument('--file', type=str)
parser.add_argument('--from', dest='date_from', type=str)
parser.add_argument('--to',   dest='date_to',   type=str)
args = parser.parse_args()

print("\n" + "="*60)
print("  IBKR WHEEL STRATEGY REPORT")
print("="*60)

if args.file:
    filepath = args.file
else:
    d = os.path.dirname(os.path.abspath(__file__))
    csvfiles = sorted([f for f in os.listdir(d) if f.endswith('.csv') and 'U1807' in f])
    default = os.path.join(d, csvfiles[-1]) if csvfiles else ''
    filepath = input(f"\nCSV file [{default}]: ").strip() or default

if not os.path.exists(filepath):
    print(f"ERROR: {filepath}"); sys.exit(1)
print(f"Loading: {filepath}")

with open(filepath, 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()

section_lines = {}
for line in lines:
    parts = line.strip().split(',', 2)
    if len(parts) < 2: continue
    sec = parts[0].strip('"'); rt = parts[1].strip('"')
    if sec not in section_lines:
        section_lines[sec] = {'header':None,'data':[]}
    if rt=='Header': section_lines[sec]['header']=line
    elif rt=='Data': section_lines[sec]['data'].append(line)

sections = {}
for sec, c in section_lines.items():
    if c['header'] and c['data']:
        df = pd.read_csv(io.StringIO(c['header']+''.join(c['data'])))
        df = df.iloc[:,2:]
        sections[sec] = df

trades_all = sections['Trades'].copy()
for col in ['Quantity','T. Price','Proceeds','Comm in USD']:
    trades_all[col] = pd.to_numeric(trades_all[col], errors='coerce')
trades_all['TradeDate'] = pd.to_datetime(trades_all['Date/Time'], errors='coerce')

mn = trades_all['TradeDate'].min().strftime('%Y-%m-%d')
mx = trades_all['TradeDate'].max().strftime('%Y-%m-%d')
print(f"Data: {mn} to {mx}")

date_from = args.date_from or input(f"Start date [{mn}]: ").strip() or mn
date_to   = args.date_to   or input(f"End date   [{mx}]: ").strip() or mx
start_dt  = datetime.strptime(date_from,'%Y-%m-%d')
end_dt    = datetime.strptime(date_to,'%Y-%m-%d')
total_days= (end_dt-start_dt).days
period_label = f"{start_dt.strftime('%d %b %Y')} - {end_dt.strftime('%d %b %Y')}"
print(f"Period: {period_label}")

# Use end of day for end date so all trades on end date are included
end_dt_eod = pd.Timestamp(end_dt) + pd.Timedelta(hours=23, minutes=59, seconds=59)
trades = trades_all[(trades_all['TradeDate']>=start_dt)&(trades_all['TradeDate']<=end_dt_eod)].copy()
opt = trades[trades['Asset Category']=='Equity and Index Options'].copy()
stk = trades[trades['Asset Category']=='Stocks'].copy()
stk['TradeDate'] = pd.to_datetime(stk['Date/Time'], errors='coerce')

def parse_opt(sym):
    m = re.match(r'(\w+)\s+(\d{2}\w{3}\d{2})\s+([\d.]+)\s+([CP])', str(sym))
    if m:
        return m.group(1), pd.to_datetime(m.group(2),format='%d%b%y',errors='coerce'), float(m.group(3)), 'Put' if m.group(4)=='P' else 'Call'
    return None,None,None,None

opt[['Ticker','Expiry','Strike','OptType']] = opt['Symbol'].apply(lambda s: pd.Series(parse_opt(s)))
opt['DTE']  = (opt['Expiry']-opt['TradeDate'].dt.normalize()).dt.days
opt['Date'] = opt['TradeDate'].dt.date
opt = opt.sort_values('TradeDate').reset_index(drop=True)

pos_raw  = sections.get('Open Positions', pd.DataFrame()).copy()
opt_pos  = pos_raw[pos_raw['Asset Category']=='Equity and Index Options'] if len(pos_raw)>0 else pd.DataFrame()
active_symbols = set(opt_pos['Symbol'].tolist()) if len(opt_pos)>0 else set()
stk_pos  = pos_raw[pos_raw['Asset Category']=='Stocks'] if len(pos_raw)>0 else pd.DataFrame()
for col in ['Cost Basis','Close Price','Value','Unrealized P/L']:
    if col in stk_pos.columns:
        stk_pos[col] = pd.to_numeric(stk_pos[col],errors='coerce')

pnl = sections.get('Realized & Unrealized Performance Summary', pd.DataFrame()).copy()
for col in ['Realized Total','Unrealized Total']:
    if col in pnl.columns:
        pnl[col] = pd.to_numeric(pnl[col],errors='coerce')

def get_ticker(sym):
    m = re.match(r'(\w+)',str(sym))
    return m.group(1) if m else ''

if len(pnl)>0:
    pnl['Ticker'] = pnl['Symbol'].apply(get_ticker)
    ibkr_realized   = pnl.groupby('Ticker')['Realized Total'].sum().to_dict()
    ibkr_unrealized = pnl.groupby('Ticker')['Unrealized Total'].sum().to_dict()
else:
    ibkr_realized={}; ibkr_unrealized={}

stk2 = stk.copy()
stk2['TradeDate'] = pd.to_datetime(stk2['Date/Time'],errors='coerce')
stk_buys = stk2[stk2['Quantity']>0].copy()
assigned_tickers = set(stk_buys['Symbol'].unique())

def classify_events(opt_df, stk_buys_df):
    df = opt_df.copy()
    df['Event'] = ''
    for i, row in df.iterrows():
        qty=row['Quantity']; tkr=row['Ticker']
        price=row['T. Price']; proceeds=row['Proceeds']
        tdd=row['TradeDate'].date()
        if qty < 0:
            df.at[i,'Event'] = 'Open'
        else:
            if price==0.0 and proceeds==0.0:
                same_day = stk_buys_df[(stk_buys_df['Symbol']==tkr)&(stk_buys_df['TradeDate'].dt.date==tdd)]
                df.at[i,'Event'] = 'Assigned' if len(same_day)>0 else 'Expired'
            else:
                df.at[i,'Event'] = 'Close'
    for (date,ticker),grp in df.groupby(['Date','Ticker']):
        buys  = grp[(grp['Quantity']>0)&(grp['Event']=='Close')].copy()
        sells = grp[(grp['Quantity']<0)&(grp['Event']=='Open')].copy()
        if len(buys)==0 or len(sells)==0: continue
        used_sells = set()
        for b_idx in buys.index:
            b_row=df.loc[b_idx]; b_type=b_row['OptType']
            b_sym=b_row['Symbol']; b_exp=b_row['Expiry']
            matching=sells[(sells['OptType']==b_type)&(sells['Symbol']!=b_sym)&(~sells.index.isin(used_sells))].copy()
            if len(matching)>0:
                matching['expiry_diff']=(matching['Expiry']-b_exp).abs()
                s_idx=matching.sort_values('expiry_diff').index[0]
                df.at[b_idx,'Event']='Roll/Bought'
                df.at[s_idx,'Event']='Roll/Bought'
                used_sells.add(s_idx)
    return df

opt = classify_events(opt, stk_buys)

def calc_break_even(stk_buys_df, opt_df):
    results = []
    for _,s in stk_buys_df.iterrows():
        tkr=s['Symbol']; tdd=s['TradeDate'].date()
        qty=s['Quantity']; price=s['T. Price']
        cost=abs(s['Proceeds'])+(abs(s['Comm in USD']) if pd.notna(s['Comm in USD']) else 0)
        zero_buy = opt_df[(opt_df['Ticker']==tkr)&(opt_df['Quantity']>0)&
            (opt_df['TradeDate'].dt.date==tdd)&(opt_df['T. Price']==0.0)&(opt_df['Proceeds']==0.0)]
        put_net=0.0; put_sym='Market Purchase'
        if len(zero_buy)>0:
            assigned_sym = zero_buy.iloc[0]['Symbol']
            orig_sell = opt_df[(opt_df['Symbol']==assigned_sym)&(opt_df['Quantity']<0)]
            if len(orig_sell)>0:
                p=orig_sell.iloc[-1]
                c=p['Comm in USD'] if pd.notna(p['Comm in USD']) else 0
                put_net=p['Proceeds']+c; put_sym=assigned_sym
        be=(cost-put_net)/qty
        results.append({'Date':tdd,'Ticker':tkr,'Qty':int(qty),'Price':price,
            'Cost':cost,'Assigning Put':put_sym,'Put Net':put_net,
            'Adj Cost':round(cost-put_net,2),'Break Even':round(be,2)})
    return results

be_results = calc_break_even(stk_buys, opt)
print(f"Break evens: {[(r['Ticker'],r['Break Even']) for r in be_results]}")

gross    = opt[opt['Quantity']<0]['Proceeds'].sum()
paid     = opt[opt['Quantity']>0]['Proceeds'].sum()
comm_opt = opt['Comm in USD'].sum()
true_net = gross+paid+comm_opt
stock_cost = stk_pos['Cost Basis'].sum() if len(stk_pos)>0 and 'Cost Basis' in stk_pos.columns else 0
period_roi = true_net/stock_cost*100 if stock_cost!=0 else 0
ann_roi    = period_roi*365/total_days if total_days>0 else 0

opt['MonthPeriod'] = opt['TradeDate'].dt.to_period('M')
mo_coll = opt[opt['Quantity']<0].groupby('MonthPeriod')['Proceeds'].sum()
mo_paid = opt[opt['Quantity']>0].groupby('MonthPeriod')['Proceeds'].sum()
mo_comm = opt.groupby('MonthPeriod')['Comm in USD'].sum()
mo_cnt  = opt[opt['Quantity']<0].groupby('MonthPeriod')['Quantity'].count()
mo = pd.DataFrame({'cnt':mo_cnt,'coll':mo_coll,'paid':mo_paid,'comm':mo_comm}).fillna(0).reset_index()
mo['net']=mo['coll']+mo['paid']+mo['comm']; mo['cum']=mo['net'].cumsum()
mo['MonthStr']=mo['MonthPeriod'].astype(str)

cap_opts = opt[opt['Quantity']<0].groupby('Ticker')['Strike'].max()*100
cap_stk  = stk_pos.set_index('Symbol')['Cost Basis'] if len(stk_pos)>0 and 'Cost Basis' in stk_pos.columns else pd.Series()
net_tkr  = opt.groupby('Ticker',group_keys=False).apply(lambda g: g['Proceeds'].sum()+g['Comm in USD'].sum())
act_days = opt.groupby('Ticker',group_keys=False).apply(lambda g: (g['TradeDate'].max()-g['TradeDate'].min()).days+1)
dte_days = opt[opt['Quantity']<0].groupby('Ticker')['DTE'].sum()
ticker_rows=[]
for tkr in sorted(net_tkr.index):
    net=net_tkr[tkr]
    if tkr in cap_stk.index and pd.notna(cap_stk[tkr]):
        capital=float(cap_stk[tkr]); cap_type='Stock Cost Basis'
    elif tkr in cap_opts.index:
        capital=float(cap_opts[tkr]); cap_type='Cash Secured'
    else:
        capital=None; cap_type='N/A'
    ad=int(act_days[tkr]); dtd=int(dte_days.get(tkr,0))
    p_roi=(net/capital*100) if capital else None
    a_act=(net/capital*365/max(ad,1)*100) if capital else None
    a_dte=(net/capital*365/max(dtd,1)*100) if capital and dtd>0 else None
    ticker_rows.append({'Ticker':tkr,'Capital ($)':capital,'Capital Type':cap_type,'Net ($)':net,
        'Active Days':ad,'DTE Days':dtd,'Period ROI (%)':p_roi,
        'Ann ROI Active (%)':a_act,'Ann ROI DTE (%)':a_dte})
rdf = pd.DataFrame(ticker_rows).sort_values('Ann ROI Active (%)',ascending=False,na_position='last')

dep2 = sections.get('Deposits & Withdrawals',pd.DataFrame()).copy()
dep2['Amount'] = pd.to_numeric(dep2.get('Amount',pd.Series()),errors='coerce')
dep2['Settle Date'] = pd.to_datetime(dep2.get('Settle Date',pd.Series()),errors='coerce')
dep2['Currency'] = dep2.get('Currency',pd.Series()).fillna('').astype(str)
dep2['Description'] = dep2.get('Description',pd.Series()).fillna('')
usd_row = dep2[dep2['Currency'].str.strip()=='Total in USD']
if len(usd_row)>0:
    deposits=float(usd_row['Amount'].iloc[0])
else:
    usd_row2=dep2[dep2['Description'].str.contains('Total in USD',case=False,na=False)]
    deposits=float(usd_row2['Amount'].iloc[0]) if len(usd_row2)>0 else 0.0
withdrawals=float(dep2[dep2['Amount']<0]['Amount'].sum()) if len(dep2)>0 else 0.0
# Parse Cash Report directly from raw CSV lines (USD rows only)
def get_cr_val(lines, label, currency='USD'):
    for line in lines:
        parts = [p.strip().strip('"') for p in line.strip().split(',')]
        if len(parts) >= 5 and parts[0]=='Cash Report' and parts[1]=='Data':
            if parts[2]==label and parts[3]==currency:
                try: return float(parts[4])
                except: pass
    return 0.0

int_total    = get_cr_val(lines, 'Broker Interest Paid and Received')
fees_total   = get_cr_val(lines, 'Other Fees')
sales_tax    = get_cr_val(lines, 'Sales Tax')
withholding  = get_cr_val(lines, 'Withholding Tax')
dividends    = get_cr_val(lines, 'Dividends')
fx_gain_loss = get_cr_val(lines, 'Cash FX Translation Gain/Loss')
ibkr_end_cash= get_cr_val(lines, 'Ending Cash')
print(f"Interest (net): ${int_total:,.2f}")
print(f"Fees: ${fees_total:,.2f}")
print(f"Sales Tax: ${sales_tax:,.2f}")
print(f"Withholding: ${withholding:,.2f}")
print(f"Dividends: ${dividends:,.2f}")
print(f"FX G/L: ${fx_gain_loss:,.2f}")
print(f"IBKR Ending Cash: ${ibkr_end_cash:,.2f}")
stk_bought=float(stk2[stk2['Quantity']>0]['Proceeds'].sum()) if len(stk2)>0 else 0.0
stk_sold  =float(stk2[stk2['Quantity']<0]['Proceeds'].sum()) if len(stk2)>0 else 0.0
stk_comm  =float(stk2['Comm in USD'].sum()) if len(stk2)>0 else 0.0
print(f"Cash deposited: ${deposits:,.2f}")

dep_clean=dep2[(dep2['Amount']>0)&(dep2['Settle Date'].notna())].copy().sort_values('Settle Date')
dep_clean['DepDate']=dep_clean['Settle Date'].dt.date
fx_all=trades_all[trades_all['Asset Category']=='Forex'].copy()
fx_all['TradeDate']=pd.to_datetime(fx_all['Date/Time'],errors='coerce')
for col in ['T. Price','Proceeds','Comm in USD']:
    fx_all[col]=pd.to_numeric(fx_all[col],errors='coerce')
fx_usd=fx_all[(fx_all['Symbol']=='USD.SGD')&(fx_all['Proceeds'].abs()>100)].copy()
fx_usd['TradeDate_date']=fx_usd['TradeDate'].dt.date
fx_usd['SGD_bought']=fx_usd['Proceeds'].abs()*fx_usd['T. Price']
sgd_rows=[]; used_fx=set()
for _,d in dep_clean.iterrows():
    dep_date=d['DepDate']; sgd_amt=d['Amount']
    cands=fx_usd[(fx_usd['TradeDate_date']>=dep_date-pd.Timedelta(days=2))&
        (fx_usd['TradeDate_date']<=dep_date+pd.Timedelta(days=2))&
        (~fx_usd.index.isin(used_fx))].copy()
    if len(cands)>0:
        cands['diff']=(cands['SGD_bought']-sgd_amt).abs()
        best=cands.sort_values('diff').iloc[0]; used_fx.add(cands.sort_values('diff').index[0])
        usd_recv=round(sgd_amt/best['T. Price'],2)
        comm_fx=round(best['Comm in USD'],2) if pd.notna(best['Comm in USD']) else 0.0
        sgd_rows.append({'Settle Date':d['Settle Date'].strftime('%d %b %Y'),
            'SGD Deposited':sgd_amt,'Exchange Rate':round(best['T. Price'],5),
            'USD Received':usd_recv,'Commission (USD)':comm_fx,
            'Net USD':round(usd_recv+comm_fx,2),'Forex Date':best['TradeDate'].strftime('%d %b %Y')})
    else:
        avg=fx_usd['T. Price'].mean() if len(fx_usd)>0 else 1.29
        sgd_rows.append({'Settle Date':d['Settle Date'].strftime('%d %b %Y'),
            'SGD Deposited':sgd_amt,'Exchange Rate':round(avg,5),
            'USD Received':round(sgd_amt/avg,2),'Commission (USD)':0.0,
            'Net USD':round(sgd_amt/avg,2),'Forex Date':'Est.'})
sgd_df=pd.DataFrame(sgd_rows) if sgd_rows else pd.DataFrame()

DARK='1F3864'; MID='2E75B6'; LITE='D6E4F7'
GL='E2EFDA'; RL='FFE0E0'; W='FFFFFF'; GR='F2F2F2'; AMBER='FFF2CC'
GD='375623'; RD='C00000'; PURPLE='E8E8FF'; BLUE='D6E4F7'

def h(cell,bg=DARK,fg=W,sz=11,wrap=False,align='center'):
    cell.font=Font(name='Calibri',bold=True,color=fg,size=sz)
    cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)

def b(cell):
    s=Side(style='thin',color='BFBFBF')
    cell.border=Border(left=s,right=s,top=s,bottom=s)

def cw(ws,col,width):
    ws.column_dimensions[get_column_letter(col)].width=width

def mc(cell,val,fmt='#,##0.00'):
    cell.number_format=fmt
    cell.font=Font(name='Calibri',size=10,
        color=GD if isinstance(val,(int,float)) and val>0 else
              RD if isinstance(val,(int,float)) and val<0 else '000000')

def plain(cell,sz=10):
    cell.font=Font(name='Calibri',size=sz)

def add_totals(ws,row,cols,label='TOTALS'):
    c=ws.cell(row=row,column=1,value=label)
    c.font=Font(bold=True,name='Calibri',color=W,size=10)
    c.fill=PatternFill('solid',start_color=DARK)
    c.alignment=Alignment(horizontal='center',vertical='center'); b(c)
    for col,letter in cols:
        cell=ws.cell(row=row,column=col,value=f'=SUM({letter}3:{letter}{row-2})')
        cell.number_format='#,##0.00'; cell.font=Font(bold=True,name='Calibri',color=W)
        cell.fill=PatternFill('solid',start_color=DARK)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)

EV_COLORS={'Open':GL,'Roll/Bought':AMBER,'Close':RL,'Expired':PURPLE,'Assigned':BLUE}
wb=Workbook()

ws1=wb.active; ws1.title='Options Trade Log'; ws1.freeze_panes='A3'
ws1.merge_cells('A1:M1')
ws1['A1']=f'Options Trade Log  |  {period_label}  |  Wheel Strategy'
h(ws1['A1'],sz=13); ws1.row_dimensions[1].height=28
for c,v in enumerate(['Date/Time','Symbol','Ticker','Expiry','Strike','Type','DTE',
    'Event','Quantity','T. Price','Proceeds','Comm in USD','MTM in USD'],1):
    cell=ws1.cell(row=2,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
ws1.row_dimensions[2].height=32
for ri,(_,row) in enumerate(opt.sort_values('TradeDate').iterrows(),3):
    event=row.get('Event',''); ebg=EV_COLORS.get(event,W); rbg=W if ri%2==0 else GR
    mtm=row.get('MTM in USD',0)
    vals=[row['TradeDate'].strftime('%d %b %Y %H:%M') if pd.notna(row['TradeDate']) else '',
        row['Symbol'],row['Ticker'],
        row['Expiry'].strftime('%d %b %Y') if pd.notna(row['Expiry']) else '',
        row['Strike'],row['OptType'],
        int(row['DTE']) if pd.notna(row['DTE']) else '',
        event,int(row['Quantity']) if pd.notna(row['Quantity']) else '',
        row['T. Price'],row['Proceeds'],row['Comm in USD'],
        mtm if pd.notna(mtm) else 0]
    for ci,val in enumerate(vals,1):
        cell=ws1.cell(row=ri,column=ci,value=val)
        cell.fill=PatternFill('solid',start_color=ebg if ci==8 else rbg)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci in [5,10,11,12,13]: mc(cell,val)
        else: plain(cell)
ldr1=2+len(opt); add_totals(ws1,ldr1+2,[(11,'K'),(12,'L')])
lr1=ldr1+4; ws1.merge_cells(f'A{lr1}:M{lr1}')
h(ws1.cell(row=lr1,column=1,value='EVENT LEGEND'),bg=DARK,sz=10)
for i,(label,color) in enumerate([
    ('Open - Sell to Open new position',GL),
    ('Roll/Bought - Buy+Sell same day same ticker',AMBER),
    ('Close - Buy to Close before expiry',RL),
    ('Expired - Expired worthless at expiry',PURPLE),
    ('Assigned - Put assigned, shares received',BLUE)],lr1+1):
    cell=ws1.cell(row=i,column=1,value=label)
    cell.fill=PatternFill('solid',start_color=color)
    cell.font=Font(name='Calibri',size=10,bold=True); b(cell)
    ws1.merge_cells(f'A{i}:M{i}')
for i,ww in enumerate([18,28,8,12,10,6,6,14,8,10,12,14,12],1): cw(ws1,i,ww)

ws2=wb.create_sheet('Ticker Summary'); ws2.freeze_panes='A3'
ws2.merge_cells('A1:K1')
ws2['A1']=f'Options P&L by Ticker  |  {period_label}'
h(ws2['A1'],sz=13); ws2.row_dimensions[1].height=28
for c,v in enumerate(['Ticker','Type','# Opens','Gross Collected ($)','Premium Paid ($)',
    'Commissions ($)','Net Premium ($)','IBKR Realized ($)','IBKR Unrealized ($)',
    'IBKR Total ($)','Status'],1):
    cell=ws2.cell(row=2,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
ws2.row_dimensions[2].height=36
sg=opt[opt['Quantity']<0].groupby(['Ticker','OptType']).agg(opens=('Quantity','count'),gross=('Proceeds','sum')).reset_index()
bg2=opt[opt['Quantity']>0].groupby(['Ticker','OptType']).agg(paid=('Proceeds','sum')).reset_index()
cg=opt.groupby(['Ticker','OptType'])['Comm in USD'].sum().reset_index(); cg.columns=['Ticker','OptType','comm']
tg=sg.merge(bg2,on=['Ticker','OptType'],how='left').merge(cg,on=['Ticker','OptType'],how='left').fillna(0)
tg['net']=tg['gross']+tg['paid']+tg['comm']
held_tickers_set = set(stk_pos['Symbol'].tolist()) if len(stk_pos)>0 else set()
for ri,(_,row) in enumerate(tg.sort_values(['Ticker','OptType']).iterrows(),3):
    tkr=row['Ticker']; otype=row['OptType']
    ibkr_r=ibkr_realized.get(tkr,0); ibkr_u=ibkr_unrealized.get(tkr,0)
    has_open_opt = any(s.startswith(tkr+' ') for s in active_symbols)
    has_stock    = tkr in held_tickers_set
    if has_stock:
        status = 'Assigned'
    elif has_open_opt:
        status = 'Active'
    else:
        status = 'Closed'
    rbg=W if ri%2==0 else GR
    for ci,val in enumerate([tkr,otype,int(row['opens']),row['gross'],row['paid'],
        row['comm'],row['net'],ibkr_r,ibkr_u,ibkr_r+ibkr_u,status],1):
        cell=ws2.cell(row=ri,column=ci,value=val)
        bg=rbg
        if ci in [7,8]: bg=GL if isinstance(val,float) and val>0 else RL
        cell.fill=PatternFill('solid',start_color=bg)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci in [4,5,6,7,8,9,10]: mc(cell,val)
        else: plain(cell)
ldr2=2+len(tg); add_totals(ws2,ldr2+2,[(4,'D'),(5,'E'),(6,'F'),(7,'G'),(8,'H'),(9,'I'),(10,'J')])
for i,ww in enumerate([10,8,8,18,16,14,16,18,18,16,10],1): cw(ws2,i,ww)

ws3=wb.create_sheet('Monthly Premium'); ws3.freeze_panes='A3'
ws3.merge_cells('A1:G1')
ws3['A1']=f'Monthly Premium Income  |  {period_label}'
h(ws3['A1'],sz=13); ws3.row_dimensions[1].height=28
for c,v in enumerate(['Month','# Contracts','Gross Collected ($)','Premium Paid ($)',
    'Commissions ($)','Net Premium ($)','Cumulative ($)'],1):
    cell=ws3.cell(row=2,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
ws3.row_dimensions[2].height=30
for ri,(_,row) in enumerate(mo.iterrows(),3):
    rbg=W if ri%2==0 else GR
    for ci,val in enumerate([row['MonthStr'],int(row['cnt']),row['coll'],
        row['paid'],row['comm'],row['net'],row['cum']],1):
        cell=ws3.cell(row=ri,column=ci,value=val)
        cell.fill=PatternFill('solid',start_color=rbg)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci>2: mc(cell,val)
        else: plain(cell)
ldr3=2+len(mo); add_totals(ws3,ldr3+2,[(3,'C'),(4,'D'),(5,'E'),(6,'F')])
for i,ww in enumerate([14,12,20,18,16,16,18],1): cw(ws3,i,ww)

ws4=wb.create_sheet('Cash Report')
ws4.merge_cells('A1:D1')
ws4['A1']=f'Cash Report  |  {period_label}'
h(ws4['A1'],sz=13); ws4.row_dimensions[1].height=28
def sec_hdr(ws,row,label):
    ws.merge_cells(f'A{row}:D{row}')
    cell=ws.cell(row=row,column=1,value=label); h(cell,bg=MID,sz=11,align='left')
def data_row(ws,row,label,val,is_total=False,note=''):
    lc=ws.cell(row=row,column=1,value=label); vc=ws.cell(row=row,column=2,value=val)
    nc=ws.cell(row=row,column=3,value=note)
    bg=DARK if is_total else (GR if row%2==0 else W); fg=W if is_total else '000000'
    lc.font=Font(name='Calibri',size=11,bold=is_total,color=fg)
    lc.fill=PatternFill('solid',start_color=bg)
    lc.alignment=Alignment(horizontal='left',vertical='center')
    vc.number_format='#,##0.00'
    vc.font=Font(name='Calibri',size=11,bold=is_total,
        color=W if is_total else (GD if isinstance(val,(int,float)) and val>0 else RD if isinstance(val,(int,float)) and val<0 else '000000'))
    vc.fill=PatternFill('solid',start_color=bg)
    vc.alignment=Alignment(horizontal='right',vertical='center')
    nc.font=Font(name='Calibri',size=9,italic=True,color='7F7F7F')
    nc.alignment=Alignment(horizontal='left',vertical='center')
    b(lc); b(vc)
cr=3
sec_hdr(ws4,cr,'CASH INFLOWS (USD)'); cr+=1
data_row(ws4,cr,'Cash Deposited (Wire Transfers)',deposits,note='IBKR Total in USD'); cr+=1
data_row(ws4,cr,'Options Premium Collected',gross,note='Gross from selling options'); cr+=1
data_row(ws4,cr,'Interest Credited',int_total); cr+=1
ti=deposits+gross+int_total+dividends
data_row(ws4,cr,'TOTAL INFLOWS',ti,is_total=True); cr+=2
sec_hdr(ws4,cr,'CASH OUTFLOWS (USD)'); cr+=1
data_row(ws4,cr,'Premium Paid (Buybacks & Rolls)',paid); cr+=1
data_row(ws4,cr,'Options Commissions',comm_opt); cr+=1
data_row(ws4,cr,'Stock Purchased (Assignments)',stk_bought); cr+=1
data_row(ws4,cr,'Stock Sold (Covered Calls)',stk_sold); cr+=1
data_row(ws4,cr,'Stock Commissions',stk_comm); cr+=1
if fees_total!=0: data_row(ws4,cr,'Other Fees',fees_total); cr+=1
if withdrawals!=0: data_row(ws4,cr,'Withdrawals',withdrawals); cr+=1
# Use IBKR Cash Report trade totals directly (avoids double-counting)
to=paid+comm_opt+stk_bought+stk_sold+stk_comm+fees_total+withdrawals+withholding+sales_tax
data_row(ws4,cr,'TOTAL OUTFLOWS',to,is_total=True); cr+=2
sec_hdr(ws4,cr,'SUMMARY'); cr+=1
data_row(ws4,cr,'Net Options Premium',true_net,note='Collected+paid+commissions'); cr+=1
ibkr_r_total=pnl['Realized Total'].sum() if len(pnl)>0 else 0
ibkr_u_total=pnl['Unrealized Total'].sum() if len(pnl)>0 else 0
data_row(ws4,cr,'IBKR Realized P&L (All)',ibkr_r_total,note='From IBKR P&L section'); cr+=1
data_row(ws4,cr,'IBKR Unrealized P&L (All)',ibkr_u_total); cr+=1
data_row(ws4,cr,'Interest Earned (Net)',int_total); cr+=1
data_row(ws4,cr,'Dividends Received',dividends); cr+=1
data_row(ws4,cr,'Withholding Tax',withholding); cr+=1
data_row(ws4,cr,'Sales Tax',sales_tax); cr+=1
data_row(ws4,cr,'Other Fees',fees_total); cr+=1
data_row(ws4,cr,'NET CASH FLOW',ti+to,is_total=True)
cw(ws4,1,38); cw(ws4,2,22); cw(ws4,3,40)

if len(sgd_df)>0:
    ws5=wb.create_sheet('SGD-USD Deposits'); ws5.freeze_panes='A3'
    ws5.merge_cells('A1:G1')
    ws5['A1']=f'SGD to USD Deposits  |  {period_label}'
    h(ws5['A1'],sz=13); ws5.row_dimensions[1].height=28
    for c,v in enumerate(['Settle Date','SGD Deposited','Exchange Rate (SGD/USD)',
        'USD Received','Commission (USD)','Net USD','Forex Date'],1):
        cell=ws5.cell(row=2,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
    ws5.row_dimensions[2].height=36
    for ri,(_,row) in enumerate(sgd_df.iterrows(),3):
        is_est='Est.' in str(row['Forex Date']); rbg=AMBER if is_est else (W if ri%2==0 else GR)
        for ci,val in enumerate([row['Settle Date'],row['SGD Deposited'],row['Exchange Rate'],
            row['USD Received'],row['Commission (USD)'],row['Net USD'],row['Forex Date']],1):
            cell=ws5.cell(row=ri,column=ci,value=val)
            cell.fill=PatternFill('solid',start_color=rbg)
            cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
            if ci==2: mc(cell,val)
            elif ci==3: cell.number_format='0.00000'; plain(cell)
            elif ci in [4,5,6]: mc(cell,val)
            else: plain(cell)
    ldr5=2+len(sgd_df); add_totals(ws5,ldr5+2,[(2,'B'),(4,'D'),(5,'E'),(6,'F')])
    avg_c=ws5.cell(row=ldr5+2,column=3,value=f'=AVERAGE(C3:C{ldr5})')
    avg_c.number_format='0.00000'; avg_c.font=Font(bold=True,name='Calibri',color=W)
    avg_c.fill=PatternFill('solid',start_color=DARK); b(avg_c)
    avg_c.alignment=Alignment(horizontal='center',vertical='center')
    for i,ww in enumerate([14,18,22,16,18,14,16],1): cw(ws5,i,ww)

ws6=wb.create_sheet('ROI Analysis')
ws6.merge_cells('A1:E1')
ws6['A1']=f'ROI Analysis  |  {period_label}'
h(ws6['A1'],sz=13); ws6.row_dimensions[1].height=28
sum_rows=[('PERIOD',None),('Start',start_dt.strftime('%d %b %Y')),
    ('End',end_dt.strftime('%d %b %Y')),('Days',total_days),('',''),
    ('OPTIONS INCOME',None),('Gross Collected',gross),('Premium Paid',paid),
    ('Commissions',comm_opt),('Net Premium',true_net),('',''),
    ('IBKR P&L',None),('IBKR Realized Total',ibkr_r_total),('IBKR Unrealized Total',ibkr_u_total),('',''),
    ('RETURNS',None),('Stock Cost Basis',stock_cost),
    ('Period ROI',f'{period_roi:.2f}%'),('Annualized ROI',f'{ann_roi:.2f}%')]
for ri,(label,val) in enumerate(sum_rows,3):
    lc=ws6.cell(row=ri,column=1,value=label); vc=ws6.cell(row=ri,column=2,value=val)
    is_sec=label in ('PERIOD','OPTIONS INCOME','IBKR P&L','RETURNS')
    if is_sec:
        lc.font=Font(name='Calibri',bold=True,color=W,size=11)
        lc.fill=PatternFill('solid',start_color=MID)
        ws6.merge_cells(f'A{ri}:E{ri}')
        lc.alignment=Alignment(horizontal='left',vertical='center')
    elif label:
        lc.font=Font(name='Calibri',size=11,bold=True)
        lc.fill=PatternFill('solid',start_color=GR)
        lc.alignment=Alignment(horizontal='left',vertical='center')
        vc.font=Font(name='Calibri',size=11,
            color=GD if isinstance(val,(int,float)) and val>0 else RD if isinstance(val,(int,float)) and val<0 else '000000')
        vc.fill=PatternFill('solid',start_color=W)
        vc.alignment=Alignment(horizontal='right',vertical='center')
        if isinstance(val,float): vc.number_format='#,##0.00'
    b(lc); b(vc)
tr6=len(sum_rows)+5; ws6.merge_cells(f'A{tr6}:I{tr6}')
h(ws6.cell(row=tr6,column=1,value='ROI by Ticker'),bg=DARK,sz=11); tr6+=1
for c,v in enumerate(['Ticker','Capital ($)','Capital Type','Net ($)',
    'Active Days','Ann ROI (Active Days)','DTE Days','Ann ROI (DTE Days)',
    'Period ROI (%)','# Contracts','Avg DTE'],1):
    cell=ws6.cell(row=tr6,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
ws6.row_dimensions[tr6].height=36

# Add contracts count and avg DTE per ticker
contracts_cnt = opt[opt['Quantity']<0].groupby('Ticker')['Quantity'].count()
avg_dte_per   = opt[opt['Quantity']<0].groupby('Ticker')['DTE'].mean()

for ri2,(_,row) in enumerate(rdf.iterrows(),tr6+1):
    rbg=W if ri2%2==0 else GR
    aa=row['Ann ROI Active (%)']; ad2=row['Ann ROI DTE (%)']
    tkr=row['Ticker']
    n_contracts = int(contracts_cnt.get(tkr,0))
    avg_dte_val = round(avg_dte_per.get(tkr,0),1)
    def pc(v): return GL if pd.notna(v) and v>50 else (AMBER if pd.notna(v) and v>0 else (RL if pd.notna(v) and v<0 else rbg))
    for ci,val in enumerate([tkr,row['Capital ($)'],row['Capital Type'],row['Net ($)'],
        row['Active Days'],f"{aa:.2f}%" if pd.notna(aa) else 'N/A',
        row['DTE Days'],f"{ad2:.2f}%" if pd.notna(ad2) else 'N/A',
        f"{row['Period ROI (%)']:.2f}%" if pd.notna(row['Period ROI (%)']) else 'N/A',
        n_contracts, avg_dte_val],1):
        cell=ws6.cell(row=ri2,column=ci,value=val)
        cell.fill=PatternFill('solid',start_color=pc(aa) if ci==6 else pc(ad2) if ci==8 else rbg)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci in [2,4]: mc(cell,val)
        else: plain(cell)
for i,ww in enumerate([10,16,18,16,12,20,12,20,14,12,10],1): cw(ws6,i,ww)

ws7=wb.create_sheet('Stock Assignments'); ws7.freeze_panes='A3'
ws7.merge_cells('A1:L1')
ws7['A1']=f'Stock Assignments  |  {period_label}'
h(ws7['A1'],sz=13); ws7.row_dimensions[1].height=28
for c,v in enumerate(['Date','Symbol','Qty','T. Price ($)','Proceeds ($)','Comm in USD',
    'Net ($)','Assigning Put','Put Net ($)','Adj Cost ($)','Break Even ($)','Type'],1):
    cell=ws7.cell(row=2,column=c,value=v); h(cell,bg=MID,sz=10,wrap=True); b(cell)
ws7.row_dimensions[2].height=40
be_lookup={(r['Date'],r['Ticker']):r for r in be_results}
for ri,(_,row) in enumerate(stk2.sort_values('TradeDate').iterrows(),3):
    tdd=row['TradeDate'].date(); tkr=row['Symbol']
    qty=row['Quantity']; price=row['T. Price']
    proc=row['Proceeds']; comm=row['Comm in USD'] if pd.notna(row['Comm in USD']) else 0
    net=proc+comm; is_buy=qty>0
    be=be_lookup.get((tdd,tkr),{})
    put_sym  = be.get('Assigning Put','') if is_buy else ''
    put_net  = be.get('Put Net','')       if is_buy and be.get('Assigning Put','')!='Market Purchase' else ''
    adj_cost = be.get('Adj Cost','')      if is_buy else ''
    brk_even = be.get('Break Even','')    if is_buy else ''
    row_type = ('Assignment' if be.get('Assigning Put','')!='Market Purchase' else 'Market Purchase') if is_buy else 'Stock Sale'
    rbg=LITE if ri%2==0 else W
    vals=[row['TradeDate'].strftime('%d %b %Y') if pd.notna(row['TradeDate']) else '',
        tkr,int(qty),price,proc,comm,net,put_sym,put_net,adj_cost,brk_even,row_type]
    for ci,val in enumerate(vals,1):
        cell=ws7.cell(row=ri,column=ci,value=val if val!='' else '')
        if ci==11 and val!='':
            cell.fill=PatternFill('solid',start_color=GL)
            cell.font=Font(name='Calibri',size=10,bold=True,color=GD)
        else:
            cell.fill=PatternFill('solid',start_color=rbg)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci in [4,5,6,7,9,10,11] and isinstance(val,(int,float)): mc(cell,val)
        else: plain(cell)
note_row=3+len(stk2)+2; ws7.merge_cells(f'A{note_row}:L{note_row}')
h(ws7.cell(row=note_row,column=1,value='IBKR P&L BY ASSIGNED TICKER'),bg=DARK,sz=10)
note_row+=1
for c,v in enumerate(['Ticker','IBKR Realized ($)','IBKR Unrealized ($)','IBKR Total ($)'],1):
    cell=ws7.cell(row=note_row,column=c,value=v); h(cell,bg=MID,sz=10); b(cell)
note_row+=1
for tkr in sorted(assigned_tickers):
    r_val=ibkr_realized.get(tkr,0); u_val=ibkr_unrealized.get(tkr,0); t_val=r_val+u_val
    for ci,val in enumerate([tkr,r_val,u_val,t_val],1):
        cell=ws7.cell(row=note_row,column=ci,value=val)
        cell.fill=PatternFill('solid',start_color=GL if ci>1 and isinstance(val,float) and val>0 else RL if ci>1 and isinstance(val,float) and val<0 else GR)
        cell.alignment=Alignment(horizontal='center',vertical='center'); b(cell)
        if ci>1: mc(cell,val)
        else: plain(cell)
    note_row+=1
fn_row=note_row+1; ws7.merge_cells(f'A{fn_row}:L{fn_row}')
nc=ws7.cell(row=fn_row,column=1,
    value='NOTE: Break Even = (Assignment Cost - Assigning Put Net) / Shares. Market Purchases show full cost/share. IBKR P&L uses FIFO.')
nc.font=Font(name='Calibri',size=9,italic=True,color='7F7F7F')
nc.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
ws7.row_dimensions[fn_row].height=24
for i,ww in enumerate([14,10,8,12,16,14,14,28,14,16,14,14],1): cw(ws7,i,ww)

out_name=f"IBKR_Report_{start_dt.strftime('%Y%m%d')}_{end_dt.strftime('%Y%m%d')}.xlsx"
out_path=os.path.join(os.path.dirname(os.path.abspath(__file__)),out_name)
wb.save(out_path)
print(f"\n{'='*60}")
print(f"  SUCCESS: {out_path}")
print(f"  Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"{'='*60}")